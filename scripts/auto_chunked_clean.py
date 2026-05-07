"""Auto-chunked clean: split a large input file, run defensive_clean
on each chunk in an isolated subprocess, then merge the customer
bundles into one.

The May 2026 100k incident motivated this: the pipeline OOMs around
~50k accumulated rows on the production VPS. Running each chunk in
its own process gives every chunk a fresh memory baseline, so OOM
is contained to one chunk (which we can retry) rather than killing
the whole job mid-flight.

Single command::

    python -m scripts.auto_chunked_clean \\
        --input-file /path/to/WY.csv \\
        --output-dir runtime/jobs/wy-100k

By default the script auto-chunks anything ≥ 50,000 rows into chunks
of 25,000. Tune with ``--threshold-rows`` and ``--chunk-size``.

Progress goes to stdout AND to
``<output_dir>/auto_chunked_status.json`` so the Fase 2 batch UI can
poll the same state without reinventing the data model.

The script reuses ``scripts/defensive_clean.py`` (per-chunk pipeline
runner) and ``scripts/merge_customer_bundles.py`` (final merge) from
``main``. No pipeline changes.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import subprocess
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, Iterator

if __name__ == "__main__" and __package__ in (None, ""):
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd  # noqa: E402

from app.customer_bundle import (  # noqa: E402
    CLEAN_DELIVERABLE_CSV,
    CUSTOMER_BUNDLE_DIRNAME,
    HIGH_RISK_REMOVED_CSV,
    REVIEW_PROVIDER_LIMITED_CSV,
)
from scripts.merge_customer_bundles import merge_bundles  # noqa: E402


STATUS_FILE: str = "auto_chunked_status.json"
CHUNKS_SUBDIR: str = "_chunks"

# Chunk-status enum. Keep the strings stable — Fase 2 (UI) uses this
# state file as its data contract.
CHUNK_PENDING: str = "pending"
CHUNK_RUNNING: str = "running"
CHUNK_COMPLETED: str = "completed"
CHUNK_FAILED: str = "failed"

BATCH_RUNNING: str = "running"
BATCH_COMPLETED: str = "completed"
BATCH_FAILED: str = "failed"
BATCH_PARTIAL_FAILURE: str = "partial_failure"


@dataclass
class ChunkState:
    index: int
    input_path: str
    run_dir: str
    status: str = CHUNK_PENDING
    started_at: str | None = None
    completed_at: str | None = None
    exit_code: int | None = None
    counts: dict | None = None
    error: str | None = None


@dataclass
class BatchState:
    started_at: str
    input_file: str
    input_format: str
    total_rows: int
    threshold_rows: int
    chunk_size: int
    status: str = BATCH_RUNNING
    completed_at: str | None = None
    chunks: list[ChunkState] = field(default_factory=list)
    merged_at: str | None = None
    merged_counts: dict | None = None
    error: str | None = None

    def to_dict(self) -> dict:
        out = asdict(self)
        return out


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _utcnow() -> str:
    return datetime.now(tz=timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _emit(line: str) -> None:
    """One progress line to stdout, immediately flushed."""
    print(line, flush=True)


def _write_state(state: BatchState, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(state.to_dict(), indent=2), encoding="utf-8")
    tmp.replace(path)


def _read_state(path: Path) -> BatchState | None:
    if not path.is_file():
        return None
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None
    chunks = [ChunkState(**c) for c in raw.get("chunks", [])]
    raw_state = {k: v for k, v in raw.items() if k != "chunks"}
    return BatchState(chunks=chunks, **raw_state)


# ---------------------------------------------------------------------------
# Row counting / format detection
# ---------------------------------------------------------------------------


def _detect_format(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        return "xlsx"
    return "csv"


def _count_rows(path: Path, fmt: str) -> int:
    """Count data rows (excluding header) without loading the whole file."""
    if fmt == "csv":
        # subtract 1 for the header line
        with path.open("rb") as fh:
            n = sum(1 for _ in fh)
        return max(0, n - 1)
    # xlsx
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        # max_row counts header too — subtract 1.
        return max(0, (ws.max_row or 1) - 1)
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Splitting
# ---------------------------------------------------------------------------


def _split_csv(
    src: Path, *, chunk_size: int, out_dir: Path,
) -> list[tuple[Path, int]]:
    out_dir.mkdir(parents=True, exist_ok=True)
    base = src.stem
    paths: list[tuple[Path, int]] = []
    with src.open("r", encoding="utf-8", newline="") as fh:
        reader = csv.reader(fh)
        header = next(reader, None)
        if header is None:
            return paths
        idx = 0
        rows: list[list[str]] = []
        for row in reader:
            rows.append(row)
            if len(rows) >= chunk_size:
                idx += 1
                p = out_dir / f"{base}_part_{idx}.csv"
                with p.open("w", encoding="utf-8", newline="") as out_fh:
                    w = csv.writer(out_fh)
                    w.writerow(header)
                    w.writerows(rows)
                paths.append((p, len(rows)))
                rows = []
        if rows:
            idx += 1
            p = out_dir / f"{base}_part_{idx}.csv"
            with p.open("w", encoding="utf-8", newline="") as out_fh:
                w = csv.writer(out_fh)
                w.writerow(header)
                w.writerows(rows)
            paths.append((p, len(rows)))
    return paths


def _split_xlsx(
    src: Path, *, chunk_size: int, out_dir: Path,
) -> list[tuple[Path, int]]:
    """Split an XLSX into multiple CSV files (CSV is what
    defensive_clean reads most efficiently anyway)."""
    out_dir.mkdir(parents=True, exist_ok=True)
    # Pandas can't stream xlsx by chunk; load whole file at the
    # orchestrator level. The OOM risk lives in the pipeline below,
    # not here.
    df = pd.read_excel(src, sheet_name=0, dtype=str)
    base = src.stem
    paths: list[tuple[Path, int]] = []
    total = len(df)
    for idx, start in enumerate(range(0, total, chunk_size), start=1):
        chunk = df.iloc[start: start + chunk_size]
        p = out_dir / f"{base}_part_{idx}.csv"
        chunk.to_csv(p, index=False)
        paths.append((p, len(chunk)))
    return paths


def split_input(
    src: Path, *, chunk_size: int, out_dir: Path,
) -> list[tuple[Path, int]]:
    fmt = _detect_format(src)
    if fmt == "xlsx":
        return _split_xlsx(src, chunk_size=chunk_size, out_dir=out_dir)
    return _split_csv(src, chunk_size=chunk_size, out_dir=out_dir)


# ---------------------------------------------------------------------------
# Per-chunk subprocess
# ---------------------------------------------------------------------------


def _read_bundle_counts(run_dir: Path) -> dict | None:
    """Best-effort: read the customer_bundle/ from a completed
    defensive_clean run and return per-bucket counts."""
    bundle = run_dir / CUSTOMER_BUNDLE_DIRNAME
    if not bundle.is_dir():
        return None

    def _count(filename: str) -> int:
        p = bundle / filename
        if not p.is_file() or p.stat().st_size == 0:
            return 0
        try:
            with p.open("r", encoding="utf-8", newline="") as fh:
                return max(0, sum(1 for _ in fh) - 1)
        except OSError:
            return 0

    return {
        "clean_deliverable": _count(CLEAN_DELIVERABLE_CSV),
        "review_provider_limited": _count(REVIEW_PROVIDER_LIMITED_CSV),
        "high_risk_removed": _count(HIGH_RISK_REMOVED_CSV),
    }


def run_chunk(
    chunk_input: Path,
    chunk_run_dir: Path,
    *,
    extra_env: dict | None = None,
    cancel_event: threading.Event | None = None,
    poll_interval: float = 0.5,
) -> tuple[int, str]:
    """Run defensive_clean on ``chunk_input`` in an isolated
    subprocess. Returns ``(exit_code, captured_output)``.

    Stdout/stderr are captured for the final report but NOT echoed
    line by line (defensive_clean only prints a JSON summary at the
    end anyway). The orchestrator emits its own [i/N] events so the
    operator gets visible progress at chunk boundaries.

    When ``cancel_event`` is set mid-run, the subprocess receives
    SIGTERM, then SIGKILL after a 5s grace period. The return value
    in that case is ``(-1, "<captured>\\ncancelled")``.
    """
    env = dict(os.environ)
    env["PYTHONUNBUFFERED"] = "1"
    if extra_env:
        env.update(extra_env)

    cmd = [
        sys.executable, "-m", "scripts.defensive_clean",
        "--input-file", str(chunk_input),
        "--output-dir", str(chunk_run_dir),
    ]

    if cancel_event is None:
        # Fast path: no cancellation requested → simpler subprocess.run.
        proc = subprocess.run(
            cmd, capture_output=True, text=True, env=env, check=False,
        )
        captured = (proc.stdout or "") + (
            "\n" + proc.stderr if proc.stderr else ""
        )
        return proc.returncode, captured

    # Cancellable path: Popen + poll loop.
    popen = subprocess.Popen(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        env=env,
    )
    cancelled = False
    while True:
        try:
            stdout, stderr = popen.communicate(timeout=poll_interval)
            captured = (stdout or "") + ("\n" + stderr if stderr else "")
            if cancelled:
                captured += "\ncancelled"
                return -1, captured
            return popen.returncode, captured
        except subprocess.TimeoutExpired:
            if cancel_event.is_set() and not cancelled:
                cancelled = True
                popen.terminate()
                # Give it 5s to exit gracefully, then SIGKILL.
                try:
                    popen.wait(timeout=5.0)
                except subprocess.TimeoutExpired:
                    popen.kill()
            # Loop back to communicate() to drain pipes.


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------


@dataclass
class OrchestratorOptions:
    input_file: Path
    output_dir: Path
    chunk_size: int = 25_000
    threshold_rows: int = 50_000
    allow_partial: bool = False
    cleanup: bool = False  # default: keep chunk dirs for audit
    max_parallel: int = 1  # 1 = serial (max OOM-safety); 2-4 if VPS has RAM
    # Optional: callers (e.g. the FastAPI BatchStore) can wire a
    # threading.Event to signal cancellation. Set → orchestrator
    # aborts before the next chunk and terminates any in-flight
    # subprocess.
    cancel_event: threading.Event | None = None


def run(opts: OrchestratorOptions) -> BatchState:
    src = opts.input_file.resolve()
    out_dir = opts.output_dir.resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    state_path = out_dir / STATUS_FILE

    # Lock for serialising state-file writes when chunks run in
    # parallel. _write_state already uses tmp+replace internally, but
    # multiple threads racing on it would still produce surprising
    # ordering for human readers of the file.
    state_lock = threading.Lock()

    def write_state_locked(s: BatchState) -> None:
        with state_lock:
            _write_state(s, state_path)

    # Resume support: if a state file already exists with this
    # input_file and the same chunk_size, pick up where we left off.
    existing = _read_state(state_path)
    if existing is not None and existing.input_file == str(src):
        _emit(
            f"resuming batch from {state_path} "
            f"(status={existing.status})"
        )
        state = existing
        # Re-evaluate: anything in `running` from a prior crash → reset
        # to `pending` so we re-run it.
        for ck in state.chunks:
            if ck.status == CHUNK_RUNNING:
                ck.status = CHUNK_PENDING
                ck.started_at = None
        state.status = BATCH_RUNNING
        state.completed_at = None
    else:
        fmt = _detect_format(src)
        total_rows = _count_rows(src, fmt)

        if total_rows < opts.threshold_rows:
            _emit(
                f"auto_chunked_clean: input={src} rows={total_rows:,} "
                f"format={fmt}"
            )
            _emit(
                f"below threshold ({opts.threshold_rows:,}) — "
                f"running defensive_clean directly"
            )
            single_run = out_dir
            t0 = time.monotonic()
            code, _ = run_chunk(
                src, single_run, cancel_event=opts.cancel_event,
            )
            elapsed = time.monotonic() - t0
            if code != 0:
                _emit(f"defensive_clean failed with exit code {code}")
                state = BatchState(
                    started_at=_utcnow(),
                    input_file=str(src),
                    input_format=fmt,
                    total_rows=total_rows,
                    threshold_rows=opts.threshold_rows,
                    chunk_size=opts.chunk_size,
                    status=BATCH_FAILED,
                    completed_at=_utcnow(),
                    error=f"defensive_clean exit_code={code}",
                )
                write_state_locked(state)
                return state
            counts = _read_bundle_counts(single_run) or {}
            _emit(
                f"done: "
                f"{counts.get('clean_deliverable', 0):,} clean / "
                f"{counts.get('high_risk_removed', 0):,} removed / "
                f"{counts.get('review_provider_limited', 0):,} review "
                f"({elapsed:.0f}s)"
            )
            state = BatchState(
                started_at=_utcnow(),
                input_file=str(src),
                input_format=fmt,
                total_rows=total_rows,
                threshold_rows=opts.threshold_rows,
                chunk_size=opts.chunk_size,
                status=BATCH_COMPLETED,
                completed_at=_utcnow(),
                merged_at=_utcnow(),
                merged_counts=counts,
            )
            write_state_locked(state)
            return state

        # Chunked path.
        chunks_dir = out_dir / CHUNKS_SUBDIR
        chunks_dir.mkdir(parents=True, exist_ok=True)
        chunk_paths = split_input(
            src, chunk_size=opts.chunk_size, out_dir=chunks_dir,
        )
        n = len(chunk_paths)
        _emit(
            f"auto_chunked_clean: input={src} rows={total_rows:,} "
            f"→ {n} chunks of {opts.chunk_size:,}"
        )
        # Per-chunk split events (matches the spec's stdout format).
        for i, (p, row_count) in enumerate(chunk_paths, start=1):
            _emit(f"[{i}/{n}] split: {p.name} ({row_count:,} rows)")
        state = BatchState(
            started_at=_utcnow(),
            input_file=str(src),
            input_format=fmt,
            total_rows=total_rows,
            threshold_rows=opts.threshold_rows,
            chunk_size=opts.chunk_size,
            status=BATCH_RUNNING,
        )
        for i, (p, _row_count) in enumerate(chunk_paths, start=1):
            run_dir_i = chunks_dir / f"run_{i}"
            state.chunks.append(ChunkState(
                index=i, input_path=str(p), run_dir=str(run_dir_i),
            ))
        write_state_locked(state)

    # Drain pending chunks (serial or parallel).
    n = len(state.chunks)

    def run_one(ck: ChunkState) -> None:
        # Honor cancellation BEFORE starting the subprocess.
        if opts.cancel_event is not None and opts.cancel_event.is_set():
            ck.status = CHUNK_FAILED
            ck.started_at = ck.started_at or _utcnow()
            ck.completed_at = _utcnow()
            ck.error = "cancelled before start"
            _emit(f"[{ck.index}/{n}] cancelled before start")
            write_state_locked(state)
            return
        ck.status = CHUNK_RUNNING
        ck.started_at = _utcnow()
        write_state_locked(state)
        _emit(
            f"[{ck.index}/{n}] start: defensive_clean "
            f"{Path(ck.input_path).name}"
        )
        t0 = time.monotonic()
        code, captured = run_chunk(
            Path(ck.input_path), Path(ck.run_dir),
            cancel_event=opts.cancel_event,
        )
        elapsed = time.monotonic() - t0
        ck.exit_code = code
        ck.completed_at = _utcnow()
        if code == 0:
            counts = _read_bundle_counts(Path(ck.run_dir)) or {}
            ck.status = CHUNK_COMPLETED
            ck.counts = counts
            _emit(
                f"[{ck.index}/{n}] done:  "
                f"{counts.get('clean_deliverable', 0):,} clean / "
                f"{counts.get('high_risk_removed', 0):,} removed / "
                f"{counts.get('review_provider_limited', 0):,} review "
                f"({elapsed:.0f}s)"
            )
        else:
            ck.status = CHUNK_FAILED
            tail = "\n".join(captured.splitlines()[-20:])
            ck.error = tail[-2000:] if tail else "subprocess error"
            label = "cancelled" if (
                opts.cancel_event is not None and opts.cancel_event.is_set()
            ) else f"FAIL: exit_code={code}"
            _emit(
                f"[{ck.index}/{n}] {label}; tail:\n{tail}"
            )
        write_state_locked(state)

    pending = [ck for ck in state.chunks if ck.status != CHUNK_COMPLETED]
    if not pending:
        # Resume case: everything already done.
        for ck in state.chunks:
            _emit(
                f"[{ck.index}/{n}] skip (already completed): "
                f"{ck.counts}"
            )

    if opts.max_parallel <= 1 or len(pending) <= 1:
        # Serial path.
        for ck in pending:
            run_one(ck)
            # Strict-mode early exit: any failure not from cancellation
            # aborts the whole batch.
            if ck.status == CHUNK_FAILED and not opts.allow_partial:
                state.status = BATCH_FAILED
                state.completed_at = _utcnow()
                state.error = f"chunk {ck.index} failed; aborting"
                write_state_locked(state)
                return state
    else:
        # Parallel path.
        with ThreadPoolExecutor(max_workers=opts.max_parallel) as pool:
            futures = {pool.submit(run_one, ck): ck for ck in pending}
            for fut in as_completed(futures):
                ck = futures[fut]
                # Surface unexpected exceptions from the worker.
                exc = fut.exception()
                if exc is not None:
                    ck.status = CHUNK_FAILED
                    ck.error = f"worker_exception:{type(exc).__name__}:{exc}"
                    write_state_locked(state)
                # In parallel + strict mode, signal cancellation to
                # the rest of the pool when one chunk fails.
                if (
                    ck.status == CHUNK_FAILED
                    and not opts.allow_partial
                    and opts.cancel_event is not None
                ):
                    opts.cancel_event.set()
        # After the pool drains, decide strict vs partial.
        any_failed = any(c.status == CHUNK_FAILED for c in state.chunks)
        if any_failed and not opts.allow_partial:
            state.status = BATCH_FAILED
            state.completed_at = _utcnow()
            state.error = "one or more chunks failed; aborting"
            write_state_locked(state)
            return state

    # Merge.
    completed = [c for c in state.chunks if c.status == CHUNK_COMPLETED]
    failed = [c for c in state.chunks if c.status == CHUNK_FAILED]
    if not completed:
        state.status = BATCH_FAILED
        state.completed_at = _utcnow()
        state.error = "no chunks completed"
        write_state_locked(state)
        _emit("FAIL: no chunks to merge")
        return state

    _emit(f"merge: {len(completed)} bundles → {out_dir}/{CUSTOMER_BUNDLE_DIRNAME}")
    merged_counts = merge_bundles(
        [Path(c.run_dir) for c in completed],
        output_dir=out_dir,
    )
    state.merged_at = _utcnow()
    state.merged_counts = merged_counts
    state.completed_at = _utcnow()
    state.status = (
        BATCH_PARTIAL_FAILURE if failed else BATCH_COMPLETED
    )
    write_state_locked(state)
    _emit(
        "merge: "
        f"{merged_counts['clean_deliverable']:,} clean / "
        f"{merged_counts['review_provider_limited']:,} review / "
        f"{merged_counts['high_risk_removed']:,} removed"
    )
    if failed:
        _emit(
            f"WARN: {len(failed)} chunk(s) failed; merge contains only "
            f"completed chunks. See state file for per-chunk errors."
        )
    if opts.cleanup and state.status == BATCH_COMPLETED:
        import shutil

        chunks_dir = out_dir / CHUNKS_SUBDIR
        if chunks_dir.is_dir():
            shutil.rmtree(chunks_dir, ignore_errors=True)
            _emit(f"cleanup: removed {chunks_dir}")
    _emit(f"ok: status={state.status}")
    return state


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Auto-chunked defensive cleanup. Splits a large input "
            "into manageable chunks, runs each in an isolated "
            "subprocess, and merges the customer bundles."
        ),
    )
    parser.add_argument("--input-file", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument(
        "--chunk-size", type=int, default=25_000,
        help="Rows per chunk (default 25,000).",
    )
    parser.add_argument(
        "--threshold-rows", type=int, default=50_000,
        help=(
            "Inputs smaller than this run as one chunk (default "
            "50,000)."
        ),
    )
    parser.add_argument(
        "--max-parallel", type=int, default=1,
        help=(
            "How many chunks to run in parallel (default 1 — serial,"
            " max OOM-safety). Each parallel slot still runs in its"
            " own subprocess; the parallelism only adds threads in"
            " the orchestrator. Raise to 2-4 if the VPS has RAM."
        ),
    )
    parser.add_argument(
        "--allow-partial", action="store_true",
        help=(
            "Continue past failed chunks and merge whatever "
            "succeeded. Default: abort on first failure."
        ),
    )
    parser.add_argument(
        "--cleanup", action="store_true",
        help=(
            "Remove the per-chunk run dirs after a successful "
            "merge. Default: keep them for audit."
        ),
    )
    args = parser.parse_args(argv)
    if args.max_parallel < 1:
        parser.error("--max-parallel must be >= 1")

    opts = OrchestratorOptions(
        input_file=args.input_file,
        output_dir=args.output_dir,
        chunk_size=args.chunk_size,
        threshold_rows=args.threshold_rows,
        allow_partial=args.allow_partial,
        cleanup=args.cleanup,
        max_parallel=args.max_parallel,
    )

    state = run(opts)
    if state.status in {BATCH_COMPLETED, BATCH_PARTIAL_FAILURE}:
        return 0 if state.status == BATCH_COMPLETED else 2
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
